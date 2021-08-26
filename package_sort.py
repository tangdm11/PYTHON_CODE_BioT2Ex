def sort(inputfile, sheetname, columnname,outputfile):
    from openpyxl import load_workbook,Workbook
    workbook=load_workbook(inputfile)
    sheet = workbook[sheetname]
    """将要排序的列中的单元格的行作为键，单元格的值作为键值输入一个字典"""
    row_value={}
    for cell in sheet[columnname]:
        if isinstance(cell.value,float):
            row_value[cell.row] = cell.value
    sorted_r_v = sorted(row_value.items(),key=lambda x:x[1],reverse=True)#对字典的值进行降序排序
    srow_value = dict(sorted_r_v)#items()输出值为列表，dict进行字典转换
    """创建新的工作簿并写入符合条件的行"""
    #根据行号获取到指定行后，遍历所有单元格的值组装成一个列表，用sheet.append()写入新表
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    header_lst = []
    for cell in sheet[1]:
        header_lst.append(cell.value)
    new_sheet.append(header_lst)#将原始表的表头写入新表
    #根据字典中键值降序后对应的键列表遍历旧表格中的值输入新表格
    for row in srow_value.keys():
        data_lst = []
        for cell in sheet[row]:
            data_lst.append(cell.value)
        new_sheet.append(data_lst)
    """将新表格中的值复制到"Annotated and filtered"中"""
    for i in range(1,sheet.max_row):
        for j in range(1,sheet.max_column):
            sheet.cell(row = i,column = j).value = list(new_sheet.rows)[i-1][j-1].value
    workbook.save(outputfile)

if __name__ == "__main__":
    inputfile = "D:\\1AAA\python开发\\Bio_T2Ex\CSV\\CSV.template.xlsx"
    sheetname = "Sheet"
    columnname = "G"
    outputfile = "D:\\1AAA\python开发\\Bio_T2Ex\CSV\\CSV.template111.xlsx"
    sort(inputfile,sheetname,columnname,outputfile)
    
    