def three_color_render(inputfile,sheetname,columnname,firstcolor,secondcolor,thirdcolor,outputfile):
    from openpyxl import load_workbook
    from openpyxl.styles import colors
    from openpyxl.formatting.rule import ColorScaleRule
    workbook = load_workbook(inputfile)
    sheet = workbook[sheetname]
    color_scale_rule = ColorScaleRule(start_type="percentile",
                                  start_value=0,
                                  start_color=colors.COLOR_INDEX[firstcolor],
                                  mid_type="percentile",
                                  mid_value=50,
                                  mid_color=colors.COLOR_INDEX[secondcolor],
                                  end_type="percentile",
                                  end_value=100,
                                  end_color=colors.COLOR_INDEX[thirdcolor]
                                  )
    bcell = columnname+str(1)
    ecell = columnname+str(sheet.max_row)
    sheet.conditional_formatting.add(str(bcell)+":"+str(ecell),color_scale_rule)
    workbook.save(outputfile)

if __name__ == "__main__":
    inputfile = "D:\\1AAA\python开发\\Bio_T2Ex\CSV\\CSV.template.xlsx"
    sheetname = "Sheet"
    columnname = "G"
    firstcolor = 30
    secondcolor = 1
    thirdcolor = 29
    outputfile = "D:\\1AAA\python开发\\Bio_T2Ex\CSV\\CSV.template111.xlsx"
    three_color_render(inputfile,sheetname,columnname,firstcolor,secondcolor,thirdcolor,outputfile)

    

