def two_color_render(inputfile,sheetname,columnname,firstcolor,secondcolor,outputfile):
    from openpyxl import load_workbook
    from openpyxl.styles import colors
    from openpyxl.formatting.rule import ColorScaleRule
    workbook = load_workbook(inputfile)
    sheet = workbook[sheetname]
    color_scale_rule = ColorScaleRule(start_type="min",
                                      start_color=colors.COLOR_INDEX[firstcolor],
                                      end_type="max",
                                      end_color=colors.COLOR_INDEX[secondcolor]
                                     )
    bcell = columnname+str(1)
    ecell = columnname+str(sheet.max_row)
    sheet.conditional_formatting.add(str(bcell)+":"+str(ecell),color_scale_rule)
    workbook.save(outputfile)

if __name__ == "__main__":
    inputfile = "D:\\1AAA\python开发\\Bio_T2Ex\CSV\\CSV.template.xlsx"
    sheetname = "Sheet"
    columnname = "G"
    firstcolor = 49
    secondcolor = 47
    outputfile = "D:\\1AAA\python开发\\Bio_T2Ex\CSV\\CSV.template111.xlsx"
    two_color_render(inputfile,sheetname,columnname,firstcolor,secondcolor,outputfile)

    