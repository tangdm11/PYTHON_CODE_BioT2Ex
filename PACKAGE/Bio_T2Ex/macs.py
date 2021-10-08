def macs_to_xlsx(macsfile,xlsxfile):
    import re
    from openpyxl import Workbook, workbook
    in_fh = open(macsfile)
    workbook = Workbook()
    sheet1 = workbook.worksheets[0]
    sheet1.title = "Orginal"
    sheet2 = workbook.create_sheet("Parameters",1)#创建新的"Paremeters"表格
    sheet3 = workbook.create_sheet("Annotated and filtered",2)#创建新的"Annotated and filters"表格
    for line in in_fh:
        line = line.replace("\n","")
        line_list = line.split("\t")
        sheet1.append(line_list)#每行以列表形式输出到sheet1
        bn = re.match("#",line)#正则匹配，若不满足每一行开头为#，bn返回None
        if bn != None or line == "":#满足每一行开头为#
            line = line.replace("\n","")
            line_list = line.split("\t")
            sheet2.append(line_list)#写入wb表格
        if bn == None and line != "":#不满足每一行开头为#
            line = line.replace("\n","")
            line_list = line.split("\t")
            sheet3.append(line_list)#写入"Annotated and filtered"表格    
    in_fh.close()

   
    from openpyxl.styles import Font,Alignment
    sheets = workbook.sheetnames
    for i in range(len(sheets)):
        sheet = workbook[sheets[i]]
        for row in sheet.iter_rows():
            for cell in row:            
                """将文本格式的数字转换为浮点型"""
                if cell.value != None:
                    s = cell.value
                    try:
                        cell.value = float(s)
                    except ValueError:
                        pass
                cell.font = Font(name=u"等线",sz=12)#字体设置
                cell.alignment = Alignment(vertical = "bottom")#对齐方式
    """自动列宽"""
    dims = {}
    for i in range(len(sheets)):
        sheet = workbook[sheets[i]]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    len_cell = max([(len(line.encode("utf-8"))-len(line))/2+len(line) for line in str(cell.value).split("\n")])
                    dims[cell.column_letter] = max(dims.get(cell.column_letter,0),len_cell)#dict.get(key, default=None)
                for col,value in dims.items():
                    sheet.column_dimensions[col].width = value+6 if value < 18 else 18

    from openpyxl import Workbook
    """将要排序的列中的单元格的行作为键，单元格的值作为键值输入一个字典"""
    row_value={}
    for cell in sheet["G"]:
        if isinstance(cell.value,float):
            row_value[cell.row] = cell.value
    sorted_r_v = sorted(row_value.items(),key=lambda x:x[1],reverse=True)#对键值进行降序排序
    srow_value = dict(sorted_r_v)#items()输出值为列表，dict进行字典转换
    """创建新的工作簿并写入符合条件的行"""
    #根据行号获取到指定行后，遍历所有单元格的值组装成一个列表，用sheet.append()写入新表
    #创建和原始工作表一样的表头
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    header = sheet[1]
    header_lst = []
    for cell in header:
        header_lst.append(cell.value)
    new_sheet.append(header_lst)
    #根据字典中键值降序后对应的键列表遍历旧表格中的值输入新表格
    for row in srow_value.keys():
        data_lst = []
        for cell in sheet[row]:
            data_lst.append(cell.value)
        new_sheet.append(data_lst)
    """将新表格中的值复制到"Annotated and filtered"中"""
    for i in range(1,sheet.max_row):
        for j in range(1,sheet .max_column):
            sheet.cell(row = i,column = j).value = list(new_sheet.rows)[i-1][j-1].value

    from openpyxl.comments import Comment
    comment1 = Comment("BioT2Ex:\nChromosome name","A")
    sheet["A1"].comment = comment1
    comment2 = Comment("BioT2Ex:\nStart position of peak in the chromosome","A")
    sheet["B1"].comment = comment2
    comment3 = Comment("BioT2Ex:\nEnd position of peak in the chromosome","A")
    sheet["C1"].comment = comment3
    comment4 = Comment("BioT2Ex:\nLength of peak region","A")
    sheet["D1"].comment = comment4
    comment5 = Comment("BioT2Ex:\nAbsolute peak summit position","A")
    sheet["E1"].comment = comment5
    comment6 = Comment("BioT2Ex:\npileup height at peak summit","A")
    sheet["F1"].comment = comment6
    comment7 = Comment("BioT2Ex:\n-log10(pvalue) for the peak summit","A")
    sheet["G1"].comment = comment7
    comment8 = Comment("BioT2Ex:\nfole enrichment for this peak summit against random Poisson distribution with local lambda","A")
    sheet["H1"].comment = comment8 
    comment9 = Comment("BioT2Ex:\n-log10(qvalue) at peak summit","A")
    sheet["I1"].comment = comment9
    comment10 = Comment("BioT2Ex:\npeak name","A")
    sheet["J1"].comment = comment10

    workbook.save(xlsxfile)

if __name__ == "__main__":
    macsfile = "D:\\1AAA\\python_code\\Bio_T2Ex\\MACS\\YAP1_peaks.xls"
    xlsxfile = "D:\\1AAA\\python_code\\Bio_T2Ex\\MACS\\YAP1_peaks_modified.xlsx"
    macs_to_xlsx(macsfile,xlsxfile)

