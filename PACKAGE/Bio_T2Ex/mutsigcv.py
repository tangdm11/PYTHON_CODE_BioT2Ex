def mutsigcv(inputfile,outputfile):
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    fh = open(inputfile)
    for line in fh:
        line = line.replace("\n","")
        row = line.split("\t")
        sheet.append(row)
#workbook.save("sig_genes.xlsx")
    for row in sheet.rows:
        for cell in row:
            if cell.value != None:
                s = cell.value
                try:
                    cell.value = float(s)
                except ValueError:
                    pass
    qvalue = sheet["O"]
    pvalue = sheet["N"]
    row_lst1 = []
    row_lst2 = []
    for cell in qvalue:
        if isinstance(cell.value,float) and cell.value <= 0.1:
            row_lst1.append(cell.row)
    for cell in pvalue:
        if isinstance(cell.value,float) and cell.value <= 0.01:
            row_lst2.append(cell.row)
    row_lst = [row for row in row_lst1 if row in row_lst2]
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    header = sheet[1]
    header_lst = []
    for cell in header:
        header_lst.append(cell.value)
    new_sheet.append(header_lst)
    for row in row_lst:
        data_lst = []
        for cell in sheet[row]:
            data_lst.append(cell.value)
        new_sheet.append(data_lst)

    import mygene
    mg=mygene.MyGeneInfo()
    gene_symbol = []
    for cell in new_sheet["A"]:
        gene_symbol.append(cell.value)
    out1 = mg.querymany(gene_symbol, scopes='symbol', fields='entrezgene,ensembl.gene,name,refseq.rna', species='human')
    out2 = mg.querymany(gene_symbol, scopes='alias', fields='entrezgene,ensembl.gene,name,refseq.rna', species='human')
    out = []
    for genedic in out1:
        if "notfound" not in genedic: 
            out.append(genedic)
    for genedic in out2:
        if "notfound" not in genedic: 
            out.append(genedic)
    for cell in new_sheet["A"]:
        genesym = cell.value
        id = []
        name = []
        ensemblid = []
        refseqid = []
        for gene in out:
            if genesym == gene["query"]:
                id.append(gene["_id"])
                name.append(gene["name"])
                ensemblid.append(gene["ensembl"])
                ensembls = []
                for dic in ensemblid:
                    if isinstance(dic, list):
                        ensemblid = dic
                for dic in ensemblid:
                    ensembls.append(dic["gene"])
                refseq = gene["refseq"]["rna"]
                for i in refseq:
                    if  i[0:2] == "NM":
                        refseqid.append(i)
                new_sheet.cell(row = cell.row, column = 16).value = " | ".join(id)
                new_sheet.cell(row = cell.row, column = 17).value = " | ".join(ensembls)
                new_sheet.cell(row = cell.row, column = 18).value = " | ".join(refseqid)
                new_sheet.cell(row = cell.row, column = 19).value = " | ".join(name)
    new_sheet["P1"] = "EntrezeID"
    new_sheet["Q1"] = "EnsemblID"
    new_sheet["R1"] = "RefSeqID"
    new_sheet["S1"] = "Description"
    for cell in new_sheet["P"]:
        if cell.value != None:
            s = cell.value
            try:
                cell.value = float(s)
            except ValueError:
                pass


    webname = "https://www.genenames.org/tools/search/#!/?query="
    for cell in new_sheet["A"]:
        website = webname + cell.value
        r = cell.row
        new_sheet.cell(row =r ,column = 20).value = website
    new_sheet["T1"] = "Hugo_Symbol_website"


    from openpyxl.comments import Comment
    comment1 = Comment("MutSigCV:\nname of the gene that the mutation was in.  (can also be called 'Hugo_Symbol')","A")
    new_sheet["A1"].comment = comment1
    comment2 = Comment("MutSigCV:\nexpression level of this gene, averaged across many cell lines in the Cancer Cell Line Encylcopedia","A")
    new_sheet["B1"].comment = comment2
    comment3 = Comment("MutSigCV:\nDNA replication time of this gene, ranging approximately from 100 (very early) to 1000 (very late)","A")
    new_sheet["C1"].comment = comment3
    comment4 = Comment("MutSigCV:\nchromatin compartment of this gene, measured from HiC experment, ranging approximately from -50 (very closed) to +50 (very open)","A")
    new_sheet["D1"].comment = comment4
    comment5 = Comment("MutSigCV:\nmutation counts  of 'nonsilent' (it changes the protein sequence or splice-sites)","A")
    new_sheet["E1"].comment = comment5
    comment6 = Comment("MutSigCV:\nmutation counts  of 'silent' (it is a synonymous change)","A")
    new_sheet["F1"].comment = comment6
    comment7 = Comment("MutSigCV:\nmutation counts  of 'noncoding' (it is intronic or otherwise in a flanking noncoding region)","A")
    new_sheet["G1"].comment = comment7
    comment8 = Comment("MutSigCV:\ncoverage counts of 'nonsilent' (it changes the protein sequence or splice-sites)","A")
    new_sheet["H1"].comment = comment8
    comment9 = Comment("MutSigCV:\ncoverage counts of 'silent' (it is a synonymous change)","A")
    new_sheet["I1"].comment = comment9
    comment10 = Comment("MutSigCV:\nmutation counts  of 'noncoding' (it is intronic or otherwise in a flanking noncoding region)","A")
    new_sheet["J1"].comment = comment10
    comment11 = Comment("MutSigCV:\nnnei gives the number of neighboring genes that are pooled together to compute the background mutation rate for that gene，these genes are not necessarily adjacent on the genome, but rather they have nearby covariate values","A")
    new_sheet["K1"].comment = comment11
    comment12 = Comment("MutSigCV:\nx gives the number of mutated bases in these neigboring genes that are either silent or non-coding","A")
    new_sheet["L1"].comment = comment12
    comment13 = Comment("MutSigCV:\nX gives the total number of bases related to these neighboring genes","A")
    new_sheet["M1"].comment = comment13
    comment14 = Comment("MutSigCV:\nSignificance levels","A")
    new_sheet["N1"].comment = comment14
    comment15 = Comment("MutSigCV:\nFalse-discovery rates","A")
    new_sheet["O1"].comment = comment15
    comment16 = Comment("MutSigCV:\nthe ID of the NCBI","A")
    new_sheet["P1"].comment = comment16
    comment17 = Comment("MutSigCV:\nthe ID of the Ensembl","A")
    new_sheet["Q1"].comment = comment17
    comment18 = Comment("MutSigCV:\nthe ID of the RefSeq","A")
    new_sheet["R1"].comment = comment18
    comment19 = Comment("MutSigCV:\nthe official gene name that has been approved by the HGNC and is publicly available","A")
    new_sheet["S1"].comment = comment19
    comment20 = Comment("MutSigCV:\nthe website of the gene","A")
    new_sheet["T1"].comment = comment20



    from openpyxl.styles import Font,Alignment,Border,Side
    for row in new_sheet.iter_rows():
        for cell in row:
            cell.font = Font(name="Times New Roman",sz=12)
            cell.alignment = Alignment(vertical="bottom",horizontal="center")
    for cell in new_sheet["A"]:
        cell.alignment = Alignment(vertical="bottom",horizontal="left")
    for cell in new_sheet["Q"]:
        cell.alignment = Alignment(vertical="bottom",horizontal="left")
    for cell in new_sheet["R"]:
        cell.alignment = Alignment(vertical="bottom",horizontal="left")
    for cell in new_sheet["S"]:
        cell.alignment = Alignment(vertical="bottom",horizontal="left")
    for cell in new_sheet["T"]:
        cell.alignment = Alignment(vertical="bottom",horizontal="left")
    for cell in new_sheet[1]:
        cell.font = Font(name="Times New Roman",sz=12,b=True)
        cell.border = Border(bottom=Side(border_style="medium"))
        cell.alignment = Alignment(vertical="bottom",horizontal="center")
    from openpyxl.styles import PatternFill,colors
    yellow_fground = PatternFill(fill_type="solid",fgColor=colors.COLOR_INDEX[5])
    new_sheet["O1"].fill = yellow_fground


    from openpyxl.styles import colors
    from openpyxl.formatting.rule import ColorScaleRule
    color_scale_rule = ColorScaleRule(start_type="percentile",
                                  start_value=0,
                                  start_color=colors.COLOR_INDEX[30],
                                  mid_type="percentile",
                                  mid_value=50,
                                  mid_color=colors.COLOR_INDEX[1],
                                  end_type="percentile",
                                  end_value=100,
                                  end_color=colors.COLOR_INDEX[29]
                                  )
    bcell = "O"+str(1)
    ecell = "O"+str(sheet.max_row)
    new_sheet.conditional_formatting.add(str(bcell)+":"+str(ecell),color_scale_rule)

    
    dims = {}
    for row in new_sheet.rows:
        for cell in row:
            if cell.value:
                len_cell = max([(len(line.encode("utf-8"))-len(line))/2+len(line) for line in str(cell.value).split("\n")])
                dims[cell.column_letter] = max(dims.get(cell.column_letter,0),len_cell)
    for col,value in dims.items():
        new_sheet.column_dimensions[col].width = value+4 if value<50 else 50
    
    new_workbook.save(outputfile) 

if __name__ == "__main__":
    inputfile = "D:\\1AAA\\python_code\\Bio_T2Ex\\mutsigcv\\my_results.sig_genes.txt"
    outputfile = "D:\\1AAA\\python_code\\Bio_T2Ex\\mutsigcv\\1.xlsx"
    mutsigcv(inputfile,outputfile) 