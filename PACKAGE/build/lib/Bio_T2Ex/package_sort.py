def sort(inputfile, sheetname, columnname,outputfile):
    from openpyxl import load_workbook,Workbook
    workbook=load_workbook(inputfile)
    sheet = workbook[sheetname]
    row_value={}
    for cell in sheet[columnname]:
        if isinstance(cell.value,float):
            row_value[cell.row] = cell.value
    sorted_r_v = sorted(row_value.items(),key=lambda x:x[1],reverse=True)
    srow_value = dict(sorted_r_v)
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    header_lst = []
    for cell in sheet[1]:
        header_lst.append(cell.value)
    new_sheet.append(header_lst)
    for row in srow_value.keys():
        data_lst = []
        for cell in sheet[row]:
            data_lst.append(cell.value)
        new_sheet.append(data_lst)
    new_workbook.save(outputfile)

if __name__ == "__main__":
    inputfile = "D:\\1AAA\\python_code\\Bio_T2Ex\\CSV\\CSV.template.xlsx"
    sheetname = "Sheet"
    columnname = "G"
    outputfile = "D:\\1AAA\\python_code\\Bio_T2Ex\\CSV\\CSV.template333.xlsx"
    sort(inputfile,sheetname,columnname,outputfile)
    